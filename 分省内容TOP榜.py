import pandas as pd
#生成一个播放次数排名的列表
x=pd.read_excel(r"C:\Users\Administrator\Desktop\1 咪视界内容排行走势_日_表1.xlsx")
x1=x[x['分类']=='剔重汇总']
means=x1['播放次数'].groupby(x1['渠道省份']).sum().reset_index()
#筛选后再排名
mean1=means[(means['渠道省份']!='剔重汇总')&(means['渠道省份']!='未知')&(means['渠道省份']!='全国')]
mean2=mean1.sort_values('播放次数',ascending=False)[:10].reset_index(drop=True)
#生产一个初始dataframe用来append
d0=pd.DataFrame(columns=['省份排名',
                            '省份',
                            '日VV',
                            '排名',
                            '剧集壳名称',
                            '播放用户数',
                            '有效播放用户数',
                            '播放次数',
                            '有效播放次数',
                            '播放时长（小时）',
                            '人均播放次数',
                            '人均播放时长(小时)',
                            '次均播放时长（小时）'])
#生成一个固定行数的dataframe用来添加数据
data0=pd.DataFrame(index=range(10),columns=d0.columns)
d1=pd.DataFrame(index=[0],columns=d0.columns)
for index,row in mean2.iterrows():
    data0['省份排名']=index+1
    data0['省份']=row[0]
    data0['日VV']=row[1]
    data0['排名']=range(1,11)
    data0['剧集壳名称']=pd.DataFrame(data=x1[x1['渠道省份']==row[0]].sort_values('播放次数',ascending=False)[:10],columns=['内容名称']).values
    data0['播放用户数']=pd.DataFrame(data=x1[x1['渠道省份']==row[0]].sort_values('播放次数',ascending=False)[:10],columns=['播放用户数']).values
    data0['有效播放用户数']=pd.DataFrame(data=x1[x1['渠道省份']==row[0]].sort_values('播放次数',ascending=False)[:10],columns=['有效播放用户数']).values
    data0['播放次数']=pd.DataFrame(data=x1[x1['渠道省份']==row[0]].sort_values('播放次数',ascending=False)[:10],columns=['播放次数']).values
    data0['有效播放次数']=pd.DataFrame(data=x1[x1['渠道省份']==row[0]].sort_values('播放次数',ascending=False)[:10],columns=['有效播放次数']).values
    data0['播放时长（小时）']=pd.DataFrame(data=x1[x1['渠道省份']==row[0]].sort_values('播放次数',ascending=False)[:10],columns=['播放时长（小时）']).values
    data0['人均播放次数']=pd.DataFrame(data=x1[x1['渠道省份']==row[0]].sort_values('播放次数',ascending=False)[:10],columns=['人均播放次数']).values
    data0['人均播放时长(小时)']=pd.DataFrame(data=x1[x1['渠道省份']==row[0]].sort_values('播放次数',ascending=False)[:10],columns=['人均播放时长(小时)']).values
    data0['次均播放时长（小时）']=pd.DataFrame(data=x1[x1['渠道省份']==row[0]].sort_values('播放次数',ascending=False)[:10],columns=['次均播放时长（小时）']).values
    d0=d0.append(data0)
    d0=d0.append(d1)
#生成一个新的openpyxl表格
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
wb = Workbook()
ws = wb.active
#将dataframe转换成openpyxl表格
rows = dataframe_to_rows(d0,index=False)
for r_idx, row in enumerate(rows,1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)
#合并单元格并居中
align = Alignment(horizontal='center',vertical='center')        
for i in range(10):
    for j in ['A','B','C']:
        ws.merge_cells('{}{}:{}{}'.format(j,i*11+2,j,i*11+11))
        ws['{}{}'.format(j,i*11+2)].alignment = align
#生成一个随日期变化的文件地址
from datetime import datetime,timedelta
now=datetime.today()
yesterday=now+timedelta(-1)
m='{:0>2d}'.format(yesterday.month)+ '{:0>2d}'.format(yesterday.day)
wb.save(r"C:\Users\Administrator\Desktop\{}分省内容TOP榜.xlsx".format(m))
